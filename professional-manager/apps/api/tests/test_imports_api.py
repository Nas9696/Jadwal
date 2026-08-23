import io
import uuid

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.config import settings
from app.models import (
    CurriculumRequirement,
    Grade,
    ImportJob,
    Resource,
    Section,
    SectionOffering,
    Stage,
    Subject,
    Teacher,
    TeacherSchoolMembership,
    TeachingAssignment,
)
from conftest import FIRST_SCHOOL, FIRST_TERM, SECOND_SCHOOL, SHARED_TEACHER, TEST_TENANT

HEADERS = {"X-Tenant-ID": TEST_TENANT}


def base_url(suffix: str = "", school: str = FIRST_SCHOOL) -> str:
    return f"/api/v1/schools/{school}/imports{suffix}"


def upload_csv(client: TestClient, content: bytes, term: str | None = None, name: str = "data.csv") -> dict:
    data = {"term_id": term} if term else {}
    response = client.post(
        base_url("/upload"),
        headers=HEADERS,
        data=data,
        files={"file": (name, content, "text/csv")},
    )
    assert response.status_code == 201, response.text
    return response.json()


def map_and_validate(client: TestClient, job: dict, entity: str, columns: dict[str, str], allow_updates: bool = False) -> dict:
    mapped = client.put(
        base_url(f"/{job['id']}/mapping"),
        headers=HEADERS,
        json={"sheets": {"CSV": {"entity_type": entity, "columns": columns}}, "allow_updates": allow_updates},
    )
    assert mapped.status_code == 200, mapped.text
    validated = client.post(base_url(f"/{job['id']}/validate"), headers=HEADERS)
    assert validated.status_code == 200, validated.text
    return validated.json()


def test_arabic_headers_arbitrary_order_preview_writes_no_authoritative_data_and_commit(
    client: TestClient, session: Session
) -> None:
    before = session.scalar(select(func.count()).select_from(Teacher))
    job = upload_csv(client, "التخصص,اسم المعلم,كود المعلم\nرياضيات,أحمد الجديد,T-NEW\n".encode())
    assert job["detected_sheets"][0]["entity_type"] == "teachers"
    preview = map_and_validate(
        client,
        job,
        "teachers",
        {"التخصص": "specialty", "اسم المعلم": "teacher_name", "كود المعلم": "teacher_code"},
    )
    assert preview["status"] == "ready"
    assert preview["rows"][0]["proposed_action"] == "create"
    assert session.scalar(select(func.count()).select_from(Teacher)) == before
    committed = client.post(
        base_url(f"/{job['id']}/commit"), headers=HEADERS, json={"acknowledge_warnings": True}
    )
    assert committed.status_code == 200
    assert session.scalar(select(func.count()).select_from(Teacher)) == before + 1


def test_english_aliases_bom_semicolon_and_canonical_teacher_link(client: TestClient, session: Session) -> None:
    content = "\ufeffteacher name;employee id\nمعلم مشترك;T1\n".encode("utf-8")
    job = upload_csv(client, content)
    detection = job["detected_sheets"][0]
    assert detection["suggested_mapping"] == {
        "teacher name": "teacher_name",
        "employee id": "teacher_code",
    }
    preview = map_and_validate(
        client,
        job,
        "teachers",
        {"teacher name": "teacher_name", "employee id": "teacher_code"},
    )
    assert preview["rows"][0]["proposed_action"] == "link_existing"
    teacher_count = session.scalar(select(func.count()).select_from(Teacher))
    response = client.post(base_url(f"/{job['id']}/commit"), headers=HEADERS, json={})
    assert response.status_code == 200
    assert session.scalar(select(func.count()).select_from(Teacher)) == teacher_count
    membership = session.scalar(
        select(TeacherSchoolMembership).where(
            TeacherSchoolMembership.teacher_id == uuid.UUID(SHARED_TEACHER),
            TeacherSchoolMembership.school_id == uuid.UUID(FIRST_SCHOOL),
        )
    )
    assert membership is not None


def test_xlsx_multisheet_detection_and_formula_rejection(client: TestClient) -> None:
    workbook = Workbook()
    teachers = workbook.active
    teachers.title = "المعلمون"
    teachers.append(["كود المعلم", "اسم المعلم"])
    teachers.append(["T-X", "=CONCAT(\"أ\",\"حمد\")"])
    subjects = workbook.create_sheet("Subjects")
    subjects.append(["subject code", "subject name"])
    subjects.append(["S-X", "Science"])
    stream = io.BytesIO()
    workbook.save(stream)
    response = client.post(
        base_url("/upload"),
        headers=HEADERS,
        files={"file": ("book.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201
    job = response.json()
    assert len(job["detected_sheets"]) == 2
    mapped = client.put(
        base_url(f"/{job['id']}/mapping"),
        headers=HEADERS,
        json={"sheets": {"المعلمون": {"entity_type": "teachers", "columns": {"كود المعلم": "teacher_code", "اسم المعلم": "teacher_name"}}, "Subjects": {"entity_type": "skip", "columns": {}}}},
    )
    assert mapped.status_code == 200
    preview = client.post(base_url(f"/{job['id']}/validate"), headers=HEADERS).json()
    assert preview["status"] == "validated"
    teacher_row = next(item for item in preview["rows"] if item["sheet_name"] == "المعلمون")
    assert any(item["code"] == "formula_not_allowed" for item in teacher_row["diagnostics"])


def test_upload_security_types_and_limits(client: TestClient) -> None:
    for name in ("bad.xls", "bad.xlsm", "bad.zip", "bad.exe"):
        response = client.post(
            base_url("/upload"), headers=HEADERS, files={"file": (name, b"MZbad", "application/octet-stream")}
        )
        assert response.status_code == 415
    original = settings.import_max_file_bytes
    settings.import_max_file_bytes = 5
    try:
        limited = client.post(
            base_url("/upload"), headers=HEADERS, files={"file": ("large.csv", b"header\nvalue", "text/csv")}
        )
        assert limited.status_code == 413
    finally:
        settings.import_max_file_bytes = original
    assert client.post(base_url("/upload"), headers=HEADERS, files={"file": ("fake.csv", b"PK malicious", "text/csv")}).status_code == 415


def test_row_sheet_and_duplicate_mapping_limits(client: TestClient) -> None:
    original_rows, original_sheets = settings.import_max_rows, settings.import_max_sheets
    try:
        settings.import_max_rows = 1
        assert client.post(base_url("/upload"), headers=HEADERS, files={"file": ("rows.csv", b"a\n1\n2\n", "text/csv")}).status_code == 413
        settings.import_max_rows = original_rows
        settings.import_max_sheets = 1
        workbook = Workbook()
        workbook.active.append(["a"])
        workbook.create_sheet("second").append(["b"])
        stream = io.BytesIO()
        workbook.save(stream)
        assert client.post(base_url("/upload"), headers=HEADERS, files={"file": ("sheets.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}).status_code == 413
    finally:
        settings.import_max_rows, settings.import_max_sheets = original_rows, original_sheets
    job = upload_csv(client, "أ,ب\n1,2\n".encode())
    duplicate_mapping = client.put(base_url(f"/{job['id']}/mapping"), headers=HEADERS, json={"sheets": {"CSV": {"entity_type": "teachers", "columns": {"أ": "teacher_code", "ب": "teacher_code"}}}})
    assert duplicate_mapping.status_code == 422


def test_duplicate_source_rows_are_diagnosed(client: TestClient) -> None:
    job = upload_csv(client, "كود المعلم,اسم المعلم\nDUP,أحمد\nDUP,أحمد\n".encode())
    preview = map_and_validate(client, job, "teachers", {"كود المعلم": "teacher_code", "اسم المعلم": "teacher_name"})
    assert preview["status"] == "validated"
    assert any(item["code"] == "duplicate_row" for item in preview["rows"][1]["diagnostics"])


def test_duplicate_hash_warning_durable_result_and_double_commit(client: TestClient) -> None:
    content = "رمز المادة,اسم المادة\nNEW-S,مادة جديدة\n".encode()
    job = upload_csv(client, content)
    preview = map_and_validate(
        client,
        job,
        "subjects",
        {"رمز المادة": "subject_code", "اسم المادة": "subject_name"},
    )
    assert preview["status"] == "ready"
    first = client.post(base_url(f"/{job['id']}/commit"), headers=HEADERS, json={})
    assert first.status_code == 200
    assert client.get(base_url(f"/{job['id']}"), headers=HEADERS).json()["status"] == "committed"
    assert client.post(base_url(f"/{job['id']}/commit"), headers=HEADERS, json={}).status_code == 409
    replay = upload_csv(client, content)
    assert replay["duplicate_file_warning"] is True


def test_atomic_failure_rolls_back_all_authoritative_rows(client: TestClient, session: Session) -> None:
    content = "رمز المادة,اسم المادة\nATOMIC-S,مادة ذرية\n".encode()
    job = upload_csv(client, content)
    map_and_validate(client, job, "subjects", {"رمز المادة": "subject_code", "اسم المادة": "subject_name"})
    row = session.query(ImportJob).filter_by(id=uuid.UUID(job["id"])).one()
    staged = row.result_summary
    row.result_summary = staged
    session.commit()
    # Force an unexpected authoritative uniqueness failure after validation.
    session.add(Subject(tenant_id=uuid.UUID(TEST_TENANT), school_id=uuid.UUID(FIRST_SCHOOL), code="ATOMIC-S", name_ar="سابق"))
    session.commit()
    response = client.post(base_url(f"/{job['id']}/commit"), headers=HEADERS, json={})
    assert response.status_code == 409
    assert session.scalar(select(func.count()).select_from(Subject).where(Subject.code == "ATOMIC-S")) == 1
    assert client.get(base_url(f"/{job['id']}"), headers=HEADERS).json()["status"] == "failed"


def test_import_job_is_school_scoped_and_templates_are_safe(client: TestClient) -> None:
    job = upload_csv(client, "رمز المادة,اسم المادة\nX,س\n".encode())
    assert client.get(base_url(f"/{job['id']}", SECOND_SCHOOL), headers=HEADERS).status_code == 404
    template = client.get(base_url("/templates/assignments.csv"), headers=HEADERS)
    assert template.status_code == 200
    assert template.content.startswith(b"\xef\xbb\xbf")


def test_archived_and_ambiguous_teacher_are_conflicts(client: TestClient, session: Session) -> None:
    teacher = session.get(Teacher, uuid.UUID(SHARED_TEACHER))
    assert teacher
    teacher.is_active = False
    session.commit()
    archived = upload_csv(client, "كود المعلم,اسم المعلم\nT1,معلم مشترك\n".encode())
    preview = map_and_validate(
        client,
        archived,
        "teachers",
        {"كود المعلم": "teacher_code", "اسم المعلم": "teacher_name"},
    )
    assert preview["rows"][0]["proposed_action"] == "conflict"
    codes = [item["code"] for item in preview["rows"][0]["diagnostics"]]
    assert "teacher_archived" in codes
    ambiguous = upload_csv(client, "اسم المعلم,كود المعلم\nاسم بلا كود,\n".encode())
    ambiguous_preview = map_and_validate(
        client,
        ambiguous,
        "teachers",
        {"اسم المعلم": "teacher_name", "كود المعلم": "teacher_code"},
    )
    assert ambiguous_preview["rows"][0]["proposed_action"] == "conflict"
    assert any(item["code"] == "teacher_ambiguous" for item in ambiguous_preview["rows"][0]["diagnostics"])


def test_row_exclusion_is_respected(client: TestClient, session: Session) -> None:
    job = upload_csv(client, "كود المعلم,اسم المعلم\nE1,الأول\nE2,الثاني\n".encode())
    mapped = client.put(
        base_url(f"/{job['id']}/mapping"),
        headers=HEADERS,
        json={"sheets": {"CSV": {"entity_type": "teachers", "columns": {"كود المعلم": "teacher_code", "اسم المعلم": "teacher_name"}}}},
    ).json()
    excluded_id = mapped["rows"][1]["id"]
    assert client.put(
        base_url(f"/{job['id']}/exclude"), headers=HEADERS, json={"row_ids": [excluded_id]}
    ).status_code == 200
    preview = client.post(base_url(f"/{job['id']}/validate"), headers=HEADERS).json()
    assert preview["status"] == "ready"
    assert client.post(base_url(f"/{job['id']}/commit"), headers=HEADERS, json={}).status_code == 200
    assert session.scalar(select(func.count()).select_from(Teacher).where(Teacher.canonical_code.in_(["E1", "E2"]))) == 1


def test_wrong_school_term_is_rejected_at_upload(client: TestClient) -> None:
    response = client.post(
        base_url("/upload"),
        headers=HEADERS,
        data={"term_id": FIRST_TERM},
        files={"file": ("data.csv", b"code,name\n1,x", "text/csv")},
    )
    assert response.status_code == 201
    wrong = client.post(
        base_url("/upload", SECOND_SCHOOL),
        headers=HEADERS,
        data={"term_id": FIRST_TERM},
        files={"file": ("data.csv", b"code,name\n1,x", "text/csv")},
    )
    assert wrong.status_code == 422


def test_explicit_assignment_group_key_creates_combined_coteaching_group(client: TestClient) -> None:
    shift = client.post(
        f"/api/v1/schools/{FIRST_SCHOOL}/setup/shifts",
        headers=HEADERS,
        json={"code": "AM-I", "name_ar": "صباحي", "order": 0, "is_active": True},
    ).json()
    setup = client.get(f"/api/v1/schools/{FIRST_SCHOOL}/setup", headers=HEADERS).json()
    first_section = setup["sections"][0]
    second_section = client.post(
        f"/api/v1/schools/{FIRST_SCHOOL}/setup/sections",
        headers=HEADERS,
        json={"grade_id": first_section["grade_id"], "name_ar": "ب", "capacity": 30},
    ).json()
    for section in (first_section, second_section):
        response = client.put(
            f"/api/v1/schools/{FIRST_SCHOOL}/assignments/section-offerings",
            headers=HEADERS,
            json={"offerings": [{"term_id": FIRST_TERM, "section_id": section["id"], "shift_id": shift["id"], "is_active": True}]},
        )
        assert response.status_code == 200
    assert client.post(
        f"/api/v1/schools/{FIRST_SCHOOL}/teacher-memberships",
        headers=HEADERS,
        json={"teacher_id": SHARED_TEACHER, "is_active": True, "is_home_school": False},
    ).status_code == 201
    second_teacher = client.post(
        f"/api/v1/schools/{FIRST_SCHOOL}/teachers",
        headers=HEADERS,
        json={"canonical_code": "T2", "name_ar": "معلم ثان", "base_workload": 0, "teaching_workload_limit": 24, "is_active": True},
    )
    assert second_teacher.status_code == 201
    content = (
        "مفتاح المجموعة,عدد الحصص,الشعبة,الصف,المرحلة,المادة,كود المعلم\n"
        "G1,2,أ,الأول,ابتدائي,رياضيات,T1\n"
        "G1,2,ب,الأول,ابتدائي,رياضيات,T2\n"
    ).encode()
    job = upload_csv(client, content, FIRST_TERM)
    preview = map_and_validate(
        client,
        job,
        "assignments",
        {
            "مفتاح المجموعة": "group_key", "عدد الحصص": "weekly_occurrences",
            "الشعبة": "section_name", "الصف": "grade_name", "المرحلة": "stage_name",
            "المادة": "subject_name", "كود المعلم": "teacher_code",
        },
    )
    assert preview["status"] == "ready"
    aggregate_preview = preview["rows"][0]["after_values"]["assignment_preview"]
    assert len(aggregate_preview["coverage"]) == 2
    assert len(aggregate_preview["teacher_workloads"]) == 2
    assert all(item["delta"] == 2 for item in aggregate_preview["coverage"])
    assert all(item["delta"] == 2 for item in aggregate_preview["teacher_workloads"])
    committed = client.post(
        base_url(f"/{job['id']}/commit"), headers=HEADERS, json={"acknowledge_warnings": True}
    )
    assert committed.status_code == 200, committed.text
    snapshot = client.get(
        f"/api/v1/schools/{FIRST_SCHOOL}/assignments?term_id={FIRST_TERM}", headers=HEADERS
    ).json()
    assert len(snapshot["assignments"]) == 1
    assert len(snapshot["assignments"][0]["teacher_ids"]) == 2
    assert len(snapshot["assignments"][0]["section_offering_ids"]) == 2


def test_multisheet_staged_dependencies_preview_zero_writes_then_atomic_commit(
    client: TestClient, session: Session
) -> None:
    shift = client.post(
        f"/api/v1/schools/{FIRST_SCHOOL}/setup/shifts",
        headers=HEADERS,
        json={"code": "STAGED-AM", "name_ar": "صباحي مرحلي", "order": 7, "is_active": True},
    ).json()
    workbook = Workbook()
    sheets = {
        "Structure": (["stage", "stage code", "grade", "section", "capacity"], ["متوسطة جديدة", "NEW-ST", "أول متوسط جديد", "أ", "25"]),
        "Teachers": (["teacher code", "teacher name", "specialty"], ["STAGED-T", "معلم مرحلي", "رياضيات"]),
        "Subjects": (["subject code", "subject name"], ["STAGED-S", "رياضيات مرحلية"]),
        "Resources": (["resource code", "resource", "resource type", "capacity"], ["STAGED-R", "معمل مرحلي", "science_lab", "25"]),
        "Curriculum": (["stage", "grade", "subject", "weekly lessons"], ["متوسطة جديدة", "أول متوسط جديد", "رياضيات مرحلية", "2"]),
        "Offerings": (["stage", "grade", "section", "shift"], ["متوسطة جديدة", "أول متوسط جديد", "أ", "صباحي مرحلي"]),
        "Assignments": (["group key", "teacher code", "subject", "stage", "grade", "section", "weekly lessons", "resource code"], ["SG1", "STAGED-T", "رياضيات مرحلية", "متوسطة جديدة", "أول متوسط جديد", "أ", "2", "STAGED-R"]),
    }
    workbook.remove(workbook.active)
    for name, (headers, values) in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        sheet.append(values)
    stream = io.BytesIO()
    workbook.save(stream)
    upload = client.post(
        base_url("/upload"), headers=HEADERS, data={"term_id": FIRST_TERM},
        files={"file": ("staged.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 201, upload.text
    job = upload.json()
    mapping = {
        name: {"entity_type": entity, "columns": detected["suggested_mapping"]}
        for name, entity in (("Structure", "structure"), ("Teachers", "teachers"), ("Subjects", "subjects"), ("Resources", "resources"), ("Curriculum", "curriculum"), ("Offerings", "offerings"), ("Assignments", "assignments"))
        for detected in job["detected_sheets"] if detected["name"] == name
    }
    assert client.put(base_url(f"/{job['id']}/mapping"), headers=HEADERS, json={"sheets": mapping}).status_code == 200
    models = (Stage, Grade, Section, Teacher, Subject, Resource, CurriculumRequirement, SectionOffering, TeachingAssignment)
    before = {model: session.scalar(select(func.count()).select_from(model)) for model in models}
    preview = client.post(base_url(f"/{job['id']}/validate"), headers=HEADERS)
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["status"] == "ready", body["rows"]
    assert {model: session.scalar(select(func.count()).select_from(model)) for model in models} == before
    assignment_row = next(row for row in body["rows"] if row["entity_type"] == "assignments")
    projection = assignment_row["after_values"]["assignment_preview"]
    assert projection["coverage"][0]["required"] == 2
    assert projection["coverage"][0]["projected_assigned"] == 2
    committed = client.post(base_url(f"/{job['id']}/commit"), headers=HEADERS, json={"acknowledge_warnings": True})
    assert committed.status_code == 200, committed.text
    assert all(session.scalar(select(func.count()).select_from(model)) > before[model] for model in models)
    assert shift["id"]


def test_safe_mode_surfaces_differences_and_explicit_updates_show_before_after(
    client: TestClient, session: Session
) -> None:
    assert client.post(
        f"/api/v1/schools/{FIRST_SCHOOL}/teacher-memberships", headers=HEADERS,
        json={"teacher_id": SHARED_TEACHER, "is_active": True, "is_home_school": False},
    ).status_code == 201
    cases = [
        ("subjects", "رمز المادة,اسم المادة\nM1,رياضيات مطورة\n", {"رمز المادة": "subject_code", "اسم المادة": "subject_name"}),
        ("resources", "رمز المورد,المورد,نوع المورد,السعة\nLAB-1,معمل مطور,room,30\n", {"رمز المورد": "resource_code", "المورد": "resource_name", "نوع المورد": "resource_type", "السعة": "capacity"}),
        ("teachers", "كود المعلم,اسم المعلم,التخصص\nT1,معلم مطور,رياضيات\n", {"كود المعلم": "teacher_code", "اسم المعلم": "teacher_name", "التخصص": "specialty"}),
    ]
    for entity, text_content, columns in cases:
        safe_job = upload_csv(client, text_content.encode())
        safe = map_and_validate(client, safe_job, entity, columns)
        row = safe["rows"][0]
        assert row["proposed_action"] == "warning"
        assert row["before_values"] != row["after_values"]
        assert any(item["code"] == "existing_data_difference" for item in row["diagnostics"])
        update_job = upload_csv(client, text_content.encode(), name=f"{entity}-update.csv")
        update = map_and_validate(client, update_job, entity, columns, allow_updates=True)
        update_row = update["rows"][0]
        assert update_row["proposed_action"] == "update"
        assert update_row["before_values"] != update_row["after_values"]
        assert client.post(base_url(f"/{update_job['id']}/commit"), headers=HEADERS, json={}).status_code == 200
    assert session.scalar(select(Subject.name_ar).where(Subject.code == "M1")) == "رياضيات مطورة"
    assert session.scalar(select(Resource.name_ar).where(Resource.code == "LAB-1")) == "معمل مطور"
    assert session.get(Teacher, uuid.UUID(SHARED_TEACHER)).name_ar == "معلم مطور"


def test_same_group_key_rejects_subject_or_weekly_count_mismatch(client: TestClient) -> None:
    base = "مفتاح المجموعة,عدد الحصص,الشعبة,الصف,المرحلة,المادة,كود المعلم\n"
    for second in ("G1,5,أ,الأول,ابتدائي,علوم,T1\n", "G1,3,أ,الأول,ابتدائي,رياضيات,T1\n"):
        job = upload_csv(client, (base + "G1,5,أ,الأول,ابتدائي,رياضيات,T1\n" + second).encode(), FIRST_TERM)
        preview = map_and_validate(client, job, "assignments", {
            "مفتاح المجموعة": "group_key", "عدد الحصص": "weekly_occurrences", "الشعبة": "section_name",
            "الصف": "grade_name", "المرحلة": "stage_name", "المادة": "subject_name", "كود المعلم": "teacher_code",
        })
        assert preview["status"] == "validated"
        assert all(row["proposed_action"] == "conflict" for row in preview["rows"])
        assert all(any(item["code"] == "group_scalar_mismatch" for item in row["diagnostics"]) for row in preview["rows"])


def test_ambiguous_staged_reference_is_a_conflict(client: TestClient) -> None:
    workbook = Workbook()
    subjects = workbook.active
    subjects.title = "Subjects"
    subjects.append(["subject code", "subject name"])
    subjects.append(["AMB-1", "مادة ملتبسة"])
    subjects.append(["AMB-2", "مادة ملتبسة"])
    curriculum = workbook.create_sheet("Curriculum")
    curriculum.append(["stage", "grade", "subject", "weekly lessons"])
    curriculum.append(["ابتدائي", "الأول", "مادة ملتبسة", "3"])
    stream = io.BytesIO()
    workbook.save(stream)
    uploaded = client.post(base_url("/upload"), headers=HEADERS, files={"file": ("ambiguous.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}).json()
    mappings = {item["name"]: {"entity_type": "subjects" if item["name"] == "Subjects" else "curriculum", "columns": item["suggested_mapping"]} for item in uploaded["detected_sheets"]}
    assert client.put(base_url(f"/{uploaded['id']}/mapping"), headers=HEADERS, json={"sheets": mappings}).status_code == 200
    preview = client.post(base_url(f"/{uploaded['id']}/validate"), headers=HEADERS).json()
    curriculum_row = next(row for row in preview["rows"] if row["entity_type"] == "curriculum")
    assert curriculum_row["proposed_action"] == "conflict"
    assert any(item["code"] == "staged_reference_ambiguous" for item in curriculum_row["diagnostics"])
