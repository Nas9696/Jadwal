from __future__ import annotations

import base64
import binascii
import io
import json
import os
import uuid
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast
from xml.sax.saxutils import escape

import arabic_reshaper  # type: ignore[import-untyped]
import qrcode  # type: ignore[import-untyped]
from bidi.algorithm import get_display  # type: ignore[import-untyped]
from fastapi import HTTPException
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.drawing.image import Image as ExcelImage  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont
from reportlab.lib import colors  # type: ignore[import-untyped]
from reportlab.lib.enums import TA_RIGHT  # type: ignore[import-untyped]
from reportlab.lib.pagesizes import A3, A4, landscape, portrait  # type: ignore[import-untyped]
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # type: ignore[import-untyped]
from reportlab.pdfbase import pdfmetrics  # type: ignore[import-untyped]
from reportlab.pdfbase.ttfonts import TTFont  # type: ignore[import-untyped]
from reportlab.platypus import (  # type: ignore[import-untyped]
    Image as PdfImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.report_schemas import ReportDataset, ReportExportMetadata, ReportRow

MAX_LOGO_BYTES = 1_000_000
MAX_LOGO_PIXELS = 16_000_000
FONT_NAME = "PMArabic"
_font_registered = False


@dataclass(frozen=True)
class ExportedReport:
    content: bytes
    metadata: ReportExportMetadata


def _font_path() -> Path:
    configured = os.getenv("PM_REPORT_FONT_PATH")
    candidates = [
        Path(configured) if configured else None,
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
    ]
    for candidate in candidates:
        if candidate and candidate.is_file():
            return candidate
    raise HTTPException(500, detail={"code": "arabic_report_font_unavailable"})


def _register_font() -> Path:
    global _font_registered
    path = _font_path()
    if not _font_registered:
        pdfmetrics.registerFont(TTFont(FONT_NAME, str(path)))
        _font_registered = True
    return path


def rtl(value: Any) -> str:
    text = "" if value is None else str(value)
    return cast(str, get_display(arabic_reshaper.reshape(text)))


def _minute(value: int | None) -> str:
    if value is None:
        return ""
    return f"{value // 60:02d}:{value % 60:02d}"


def _row_values(dataset: ReportDataset, row: ReportRow) -> list[Any]:
    values: list[Any]
    if dataset.report_type == "daily_substitutions":
        values = [
            row.weekday_label,
            row.school_name,
            row.absent_teacher_name,
            row.subject_name,
            "، ".join(row.section_names),
            row.substitute_teacher_name or "غير مغطاة",
            row.coverage_status,
        ]
        if dataset.print_options.show_period_time:
            values.insert(1, f"{_minute(row.starts_at_minute)} - {_minute(row.ends_at_minute)}")
        return values
    if dataset.report_type == "waiting_workload":
        return [
            "، ".join(row.teacher_names),
            row.base_workload,
            row.teaching_workload,
            row.substitution_count,
            row.effective_limit,
            row.remaining_capacity,
            "مستثنى" if row.exempt else "مشمول",
        ]
    values = [
        row.project_cycle_week_index,
        row.weekday_label,
        row.school_name,
        row.subject_name,
        "، ".join(row.teacher_names),
        "، ".join(row.section_names),
        row.attendance_label,
    ]
    if dataset.print_options.show_period_time:
        values.insert(2, f"{_minute(row.starts_at_minute)} - {_minute(row.ends_at_minute)}")
    if dataset.print_options.show_resource:
        values.insert(len(values) - 1, "، ".join(row.resource_names))
    return values


def validate_logo(data_url: str | None) -> PILImage.Image | None:
    if not data_url:
        return None
    allowed = {"data:image/png;base64": "PNG", "data:image/jpeg;base64": "JPEG"}
    try:
        prefix, encoded = data_url.split(",", 1)
    except ValueError as exc:
        raise HTTPException(422, detail={"code": "unsafe_logo_content"}) from exc
    expected = allowed.get(prefix.casefold())
    if not expected:
        raise HTTPException(422, detail={"code": "unsafe_logo_type"})
    try:
        raw = base64.b64decode(encoded, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise HTTPException(422, detail={"code": "unsafe_logo_content"}) from exc
    if not raw or len(raw) > MAX_LOGO_BYTES:
        raise HTTPException(422, detail={"code": "unsafe_logo_size"})
    try:
        with PILImage.open(io.BytesIO(raw)) as probe:
            if probe.format != expected or probe.width * probe.height > MAX_LOGO_PIXELS:
                raise ValueError("unsafe logo dimensions or format")
            probe.verify()
        image = PILImage.open(io.BytesIO(raw))
        image.load()
    except Exception as exc:
        raise HTTPException(422, detail={"code": "unsafe_logo_content"}) from exc
    return image.convert("RGBA")


def build_qr(payload: str) -> tuple[Any, PILImage.Image]:
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=6, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    return qr, qr.make_image(fill_color="black", back_color="white").convert("RGB")


def _image_bytes(image: PILImage.Image, format_name: str = "PNG") -> io.BytesIO:
    buffer = io.BytesIO()
    image.save(buffer, format=format_name)
    buffer.seek(0)
    return buffer


def export_xlsx(dataset: ReportDataset) -> ExportedReport:
    logo = validate_logo(dataset.branding.logo_data_url)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "التقرير"
    sheet.sheet_view.rightToLeft = True
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    last_column = get_column_letter(len(dataset.columns))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = dataset.title
    sheet["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17345B")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = dataset.subtitle or ""
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = f"المصدر: {dataset.source.kind} | revision: {dataset.source.revision if dataset.source.revision is not None else 'candidate'} | تم الإنشاء: {dataset.source.generated_at.isoformat()}"
    sheet["A3"].alignment = Alignment(horizontal="center")
    if not dataset.print_options.show_heading:
        for heading_row in (1, 2, 3):
            sheet.row_dimensions[heading_row].hidden = True
    header_fill = "DCE6F7" if dataset.print_options.theme == "color" else "E5E5E5"
    thin = Side(style="thin", color="D6DCE5")
    for index, label in enumerate(dataset.columns, 1):
        cell = sheet.cell(row=4, column=index, value=label)
        cell.font = Font(name="Arial", bold=True, color="17345B")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row_index, item in enumerate(dataset.rows, 5):
        for column_index, value in enumerate(_row_values(dataset, item), 1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        sheet.row_dimensions[row_index].height = 26 if dataset.print_options.density == "comfortable" else 20
    widths = [10, 14, 17, 24, 22, 28, 25, 22, 14]
    for index in range(1, len(dataset.columns) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[index - 1]
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3 if dataset.print_options.paper == "A3" else sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = dataset.print_options.orientation
    sheet.print_title_rows = "1:4"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    if dataset.branding.footer_text:
        sheet.oddFooter.center.text = dataset.branding.footer_text
    image_row = 1
    if logo:
        raw = _image_bytes(logo)
        drawing = ExcelImage(raw)
        drawing.width, drawing.height = 60, 60
        sheet.add_image(drawing, f"{last_column}{image_row}")
    if dataset.branding.qr_payload:
        _, qr_image = build_qr(dataset.branding.qr_payload)
        raw_qr = _image_bytes(qr_image)
        drawing = ExcelImage(raw_qr)
        drawing.width, drawing.height = 72, 72
        sheet.add_image(drawing, "A1")
    metadata = workbook.create_sheet("بيانات المصدر")
    metadata.sheet_state = "hidden"
    metadata.sheet_view.rightToLeft = True
    for row in (
        ("project_id", str(dataset.source.project_id)),
        ("source_kind", dataset.source.kind),
        ("timetable_id", str(dataset.source.timetable_id or "")),
        ("candidate_id", str(dataset.source.candidate_id or "")),
        ("source_revision", dataset.source.revision),
        ("row_count", dataset.row_count),
    ):
        metadata.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()
    return ExportedReport(
        content=content,
        metadata=ReportExportMetadata(
            filename=_safe_filename(dataset, "xlsx"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            pages=1,
            source_revision=dataset.source.revision,
            multi_page=False,
        ),
    )


def _page_size(dataset: ReportDataset) -> tuple[float, float]:
    base = A3 if dataset.print_options.paper == "A3" else A4
    value = landscape(base) if dataset.print_options.orientation == "landscape" else portrait(base)
    return cast(tuple[float, float], value)


def export_pdf(dataset: ReportDataset) -> ExportedReport:
    _register_font()
    logo = validate_logo(dataset.branding.logo_data_url)
    page_size = _page_size(dataset)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=28,
        leftMargin=28,
        topMargin=34,
        bottomMargin=48,
        title=dataset.title,
        subject=f"Professional Manager source revision {dataset.source.revision}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ArabicTitle", parent=styles["Title"], fontName=FONT_NAME, fontSize=17, leading=23, alignment=TA_RIGHT, textColor=colors.HexColor("#17345B")
    )
    body_style = ParagraphStyle(
        "ArabicBody", parent=styles["BodyText"], fontName=FONT_NAME, fontSize=8.5, leading=12, alignment=TA_RIGHT
    )
    story: list[Any] = []
    heading = []
    if dataset.branding.qr_payload:
        _, qr_image = build_qr(dataset.branding.qr_payload)
        heading.append(PdfImage(_image_bytes(qr_image), width=52, height=52))
    heading.append(Paragraph(escape(rtl(dataset.title)), title_style))
    if logo:
        heading.append(PdfImage(_image_bytes(logo), width=52, height=52))
    if dataset.print_options.show_heading:
        story.append(Table([heading]))
        story.append(Paragraph(escape(rtl(dataset.subtitle or "")), body_style))
        story.append(Paragraph(escape(rtl(f"نسخة المصدر: {dataset.source.revision if dataset.source.revision is not None else 'Candidate'}")), body_style))
        story.append(Spacer(1, 10))
    data = [[Paragraph(escape(rtl(label)), body_style) for label in dataset.columns]]
    for item in dataset.rows:
        data.append([Paragraph(escape(rtl(value)), body_style) for value in _row_values(dataset, item)])
    available = page_size[0] - 56
    widths = [available / len(dataset.columns)] * len(dataset.columns)
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="RIGHT")
    header_color = colors.HexColor("#DCE6F7") if dataset.print_options.theme == "color" else colors.HexColor("#E5E5E5")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), header_color),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#17345B")),
                ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6DCE5")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFD")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5 if dataset.print_options.density == "comfortable" else 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5 if dataset.print_options.density == "comfortable" else 3),
            ]
        )
    )
    story.append(table)
    if dataset.branding.signature_labels:
        story.append(Spacer(1, 22))
        signatures = [[Paragraph(escape(rtl(f"{label}: ____________________")), body_style) for label in dataset.branding.signature_labels]]
        story.append(Table(signatures, colWidths=[available / len(signatures[0])] * len(signatures[0])))

    def footer(canvas: Any, document: Any) -> None:
        canvas.saveState()
        canvas.setFont(FONT_NAME, 8)
        text = dataset.branding.footer_text or "المدير المحترف - الجداول الذكية"
        canvas.drawRightString(page_size[0] - 28, 22, rtl(text))
        canvas.drawString(28, 22, f"{document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    content = buffer.getvalue()
    pages = max(1, content.count(b"/Type /Page") - content.count(b"/Type /Pages"))
    return ExportedReport(
        content=content,
        metadata=ReportExportMetadata(
            filename=_safe_filename(dataset, "pdf"),
            content_type="application/pdf",
            pages=pages,
            source_revision=dataset.source.revision,
            multi_page=pages > 1,
        ),
    )


def _png_dimensions(dataset: ReportDataset) -> tuple[int, int]:
    size = (1754, 2480) if dataset.print_options.paper == "A3" else (1240, 1754)
    return size[::-1] if dataset.print_options.orientation == "landscape" else size


def _draw_text(draw: ImageDraw.ImageDraw, position: tuple[int, int], value: Any, font: ImageFont.FreeTypeFont, fill: str, anchor: str = "ra") -> None:
    draw.text(position, rtl(value), font=font, fill=fill, anchor=anchor)


def _fit_text(
    draw: ImageDraw.ImageDraw,
    value: Any,
    font_path: Path,
    max_width: float,
    preferred_size: int,
) -> tuple[str, ImageFont.FreeTypeFont]:
    text = rtl(value)
    for size in range(preferred_size, 10, -1):
        font = ImageFont.truetype(str(font_path), size)
        if draw.textlength(text, font=font) <= max_width:
            return text, font
    return text, ImageFont.truetype(str(font_path), 10)


def _png_pages(dataset: ReportDataset) -> list[bytes]:
    font_path = _register_font()
    logo = validate_logo(dataset.branding.logo_data_url)
    width, height = _png_dimensions(dataset)
    margin = 54
    header_height = 190 if dataset.print_options.show_heading else 70
    row_height = 58 if dataset.print_options.density == "comfortable" else 46
    footer_height = 75
    rows_per_page = max(1, (height - header_height - footer_height - margin) // row_height - 1)
    chunks = [dataset.rows[index : index + rows_per_page] for index in range(0, len(dataset.rows), rows_per_page)] or [[]]
    body_font = ImageFont.truetype(str(font_path), 23)
    small_font = ImageFont.truetype(str(font_path), 19)
    title_font = ImageFont.truetype(str(font_path), 39)
    pages = []
    for page_number, chunk in enumerate(chunks, 1):
        background = "#FFFFFF"
        image = PILImage.new("RGB", (width, height), background)
        draw = ImageDraw.Draw(image)
        ink = "#111827"
        navy = "#17345B"
        if dataset.print_options.show_heading:
            _draw_text(draw, (width - margin, 52), dataset.title, title_font, navy)
            _draw_text(draw, (width - margin, 105), dataset.subtitle or "", body_font, "#526079")
            _draw_text(draw, (width - margin, 140), f"نسخة المصدر: {dataset.source.revision if dataset.source.revision is not None else 'Candidate'}", small_font, "#526079")
            if logo:
                logo_copy = logo.copy()
                logo_copy.thumbnail((110, 110))
                image.paste(logo_copy.convert("RGB"), (margin, 42))
            if dataset.branding.qr_payload:
                _, qr = build_qr(dataset.branding.qr_payload)
                qr = qr.resize((115, 115))
                image.paste(qr, (margin + 125 if logo else margin, 38))
        top = header_height
        column_width = (width - margin * 2) / len(dataset.columns)
        header_fill = "#DCE6F7" if dataset.print_options.theme == "color" else "#E5E5E5"
        draw.rectangle((margin, top, width - margin, top + row_height), fill=header_fill)
        for index, label in enumerate(dataset.columns):
            right = int(width - margin - index * column_width - 8)
            text, fitted = _fit_text(draw, label, font_path, column_width - 14, 19)
            draw.text((right, top + row_height // 2), text, font=fitted, fill=navy, anchor="rm")
        for row_index, item in enumerate(chunk, 1):
            y = top + row_index * row_height
            if row_index % 2 == 0:
                draw.rectangle((margin, y, width - margin, y + row_height), fill="#F8FAFD")
            for column_index, value in enumerate(_row_values(dataset, item)):
                text, fitted = _fit_text(draw, value, font_path, column_width - 14, 19)
                right = int(width - margin - column_index * column_width - 8)
                draw.text((right, y + row_height // 2), text, font=fitted, fill=ink, anchor="rm")
            draw.line((margin, y + row_height, width - margin, y + row_height), fill="#D6DCE5", width=1)
        footer_y = height - 36
        _draw_text(draw, (width - margin, footer_y), dataset.branding.footer_text or "المدير المحترف - الجداول الذكية", small_font, "#526079")
        draw.text((margin, footer_y), f"{page_number}/{len(chunks)}", font=small_font, fill="#526079", anchor="la")
        buffer = io.BytesIO()
        image.save(buffer, "PNG", optimize=True)
        pages.append(buffer.getvalue())
    return pages


def export_png(dataset: ReportDataset) -> ExportedReport:
    pages = _png_pages(dataset)
    if len(pages) == 1:
        return ExportedReport(
            content=pages[0],
            metadata=ReportExportMetadata(
                filename=_safe_filename(dataset, "png"),
                content_type="image/png",
                pages=1,
                source_revision=dataset.source.revision,
                multi_page=False,
            ),
        )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, page in enumerate(pages, 1):
            archive.writestr(f"page-{index:03d}.png", page)
        archive.writestr(
            "manifest.json",
            json.dumps(
                {
                    "pages": len(pages),
                    "source_revision": dataset.source.revision,
                    "report_type": dataset.report_type,
                },
                ensure_ascii=False,
            ),
        )
    return ExportedReport(
        content=buffer.getvalue(),
        metadata=ReportExportMetadata(
            filename=_safe_filename(dataset, "zip"),
            content_type="application/zip",
            pages=len(pages),
            source_revision=dataset.source.revision,
            multi_page=True,
        ),
    )


def _safe_filename(dataset: ReportDataset, extension: str) -> str:
    return f"professional-manager-{dataset.report_type}-{uuid.uuid4().hex[:8]}.{extension}"


def export_report(dataset: ReportDataset, format_name: str) -> ExportedReport:
    if dataset.stale:
        raise HTTPException(409, detail={"code": "stale_report_source"})
    if format_name == "xlsx":
        return export_xlsx(dataset)
    if format_name == "pdf":
        return export_pdf(dataset)
    if format_name == "png":
        return export_png(dataset)
    raise HTTPException(422, detail={"code": "unsupported_report_format"})
