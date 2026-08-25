import csv
import hashlib
import io
import re
import unicodedata
import uuid
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import PurePath
from typing import Any, NoReturn

from fastapi import HTTPException
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app.assignment_services import preview_assignment, save_assignment, save_offerings
from app.config import settings
from app.master_services import create_teacher, link_teacher, save_catalog, school, update_teacher
from app.models import (
    CurriculumRequirement,
    AcademicYear,
    Grade,
    ImportJob,
    ImportRow,
    Resource,
    SchoolShift,
    Section,
    SectionOffering,
    Stage,
    Subject,
    Teacher,
    TeacherSchoolMembership,
    TeachingAssignment,
    TeachingAssignmentSection,
    TeachingAssignmentTeacher,
    Term,
    TimetableEntry,
    WorkingTimetableEntry,
)
from app.setup_services import save_resource


ALIASES: dict[str, set[str]] = {
    "teacher_name": {"اسم المعلم", "المعلم", "teacher", "teacher name"},
    "teacher_code": {"رقم المعلم", "كود المعلم", "الكود", "teacher code", "employee id"},
    "specialty": {"التخصص", "specialty"},
    "subject_name": {"المادة", "اسم المادة", "subject", "subject name"},
    "subject_code": {"رمز المادة", "كود المادة", "subject code"},
    "stage_name": {"المرحلة", "اسم المرحلة", "stage"},
    "stage_code": {"رمز المرحلة", "stage code"},
    "grade_name": {"الصف", "اسم الصف", "grade"},
    "section_name": {"الشعبة", "اسم الشعبة", "section", "class"},
    "weekly_occurrences": {"الحصص", "عدد الحصص", "النصاب الأسبوعي", "weekly lessons"},
    "resource_name": {"المعمل", "الغرفة", "المورد", "resource", "room"},
    "resource_code": {"رمز المورد", "كود المورد", "resource code", "room code"},
    "resource_type": {"نوع المورد", "resource type"},
    "capacity": {"السعة", "capacity"},
    "shift_name": {"الفترة", "الشفت", "shift"},
    "group_key": {"مفتاح المجموعة", "مجموعة التدريس", "group key", "teaching group"},
}
ENTITY_FIELDS = {
    "teachers": {"teacher_name", "teacher_code"},
    "subjects": {"subject_name", "subject_code"},
    "structure": {"stage_name", "grade_name", "section_name"},
    "curriculum": {"grade_name", "subject_name", "weekly_occurrences"},
    "resources": {"resource_name", "resource_code"},
    "offerings": {"stage_name", "grade_name", "section_name", "shift_name"},
    "assignments": {
        "teacher_code", "subject_name", "grade_name", "section_name", "weekly_occurrences"
    },
}
REQUIRED_FIELDS = {
    "teachers": {"teacher_name", "teacher_code"},
    "subjects": {"subject_name", "subject_code"},
    "structure": {"stage_name", "grade_name", "section_name"},
    "curriculum": {"grade_name", "subject_name", "weekly_occurrences"},
    "resources": {"resource_name", "resource_code"},
    "offerings": {"grade_name", "section_name", "shift_name"},
    "assignments": {
        "teacher_code", "subject_name", "grade_name", "section_name", "weekly_occurrences"
    },
}


def fail(code: str, status: int = 422) -> NoReturn:
    raise HTTPException(status_code=status, detail={"code": code})


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    translations: dict[str | int, str | int | None] = {
        "أ": "ا", "إ": "ا", "آ": "ا", "ى": "ي", "ـ": ""
    }
    text = text.translate(str.maketrans(translations))
    return re.sub(r"[\s\-_./\\:،,;؛()\[\]]+", " ", text).strip()


NORMALIZED_ALIASES = {
    alias: target for target, aliases in ALIASES.items() for alias in map(normalize_text, aliases)
}


def _suggest_mapping(headers: list[str]) -> dict[str, str]:
    return {
        header: NORMALIZED_ALIASES[normalized]
        for header in headers
        if (normalized := normalize_text(header)) in NORMALIZED_ALIASES
    }


def _detect_entity(mapping: dict[str, str]) -> tuple[str, float]:
    found = set(mapping.values())
    scores = {
        entity: len(found & fields) / len(fields) for entity, fields in ENTITY_FIELDS.items()
    }
    entity, score = max(scores.items(), key=lambda item: item[1])
    return entity, round(score, 2)


def _safe_filename(filename: str | None) -> str:
    name = PurePath(filename or "import").name.replace("\x00", "")[:255]
    return name or "import"


def _parse_csv(content: bytes) -> list[tuple[str, list[str], list[tuple[int, dict[str, Any]]]]]:
    if content.startswith((b"PK", b"MZ")):
        fail("content_signature_mismatch", 415)
    if b"\x00" in content:
        fail("binary_file_rejected", 415)
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        fail("csv_must_be_utf8", 415)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        fail("empty_file")
    headers = [str(item).strip()[:150] for item in rows[0]]
    parsed = []
    for number, values in enumerate(rows[1:], 2):
        if not any(str(value).strip() for value in values):
            continue
        parsed.append((number, {header: str(values[index]).strip()[:1000] if index < len(values) else "" for index, header in enumerate(headers)}))
    return [("CSV", headers, parsed)]


def _inspect_xlsx(content: bytes) -> None:
    if not content.startswith(b"PK"):
        fail("xlsx_signature_mismatch", 415)
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > settings.import_max_zip_entries:
                fail("xlsx_zip_entry_limit", 413)
            if sum(item.file_size for item in entries) > settings.import_max_expanded_bytes:
                fail("xlsx_expanded_size_limit", 413)
            names = {item.filename.lower() for item in entries}
            if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
                fail("xlsx_structure_invalid", 415)
            if any("vbaproject" in name or name.endswith(".bin") for name in names):
                fail("macro_content_rejected", 415)
    except zipfile.BadZipFile:
        fail("xlsx_structure_invalid", 415)


def _parse_xlsx(content: bytes) -> list[tuple[str, list[str], list[tuple[int, dict[str, Any]]]]]:
    _inspect_xlsx(content)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    if len(workbook.sheetnames) > settings.import_max_sheets:
        fail("sheet_limit_exceeded", 413)
    result = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows()
        first = next(iterator, None)
        if first is None:
            continue
        headers = [str(cell.value or "").strip()[:150] for cell in first]
        rows = []
        for number, cells in enumerate(iterator, 2):
            values: dict[str, Any] = {}
            for index, header in enumerate(headers):
                cell = cells[index] if index < len(cells) else None
                if cell is None:
                    values[header] = ""
                elif cell.data_type == "f":
                    values[header] = {"formula": str(cell.value)[:1000]}
                else:
                    values[header] = str(cell.value or "").strip()[:1000]
            if any(value for value in values.values()):
                rows.append((number, values))
        result.append((sheet.title[:150], headers, rows))
    workbook.close()
    return result


ASC_GRADE_NAMES = {
    1: ("المرحلة الابتدائية", "الأول"),
    2: ("المرحلة الابتدائية", "الثاني"),
    3: ("المرحلة الابتدائية", "الثالث"),
    4: ("المرحلة الابتدائية", "الرابع"),
    5: ("المرحلة الابتدائية", "الخامس"),
    6: ("المرحلة الابتدائية", "السادس"),
    7: ("المرحلة المتوسطة", "الأول المتوسط"),
    8: ("المرحلة المتوسطة", "الثاني المتوسط"),
    9: ("المرحلة المتوسطة", "الثالث المتوسط"),
    10: ("المرحلة الثانوية", "الأول الثانوي"),
    11: ("المرحلة الثانوية", "الثاني الثانوي"),
    12: ("المرحلة الثانوية", "الثالث الثانوي"),
}
ASC_SHEET_ENTITIES = {
    "المعلمون من Timetables": "teachers",
    "المواد من Timetables": "subjects",
    "الصفوف والفصول من Timetables": "structure",
    "الفصول الدراسية من Timetables": "offerings",
    "الإسنادات من Timetables": "assignments",
}


def _asc_values(value: str | None) -> list[str]:
    return [item.strip() for item in re.split(r"[,;]", value or "") if item.strip()]


def _asc_class_name(value: str) -> tuple[str, str, str] | None:
    match = re.match(r"^\s*([0-9٠-٩]+)\s*[/\\-]\s*(.+?)\s*$", value)
    if not match:
        return None
    digits = match.group(1).translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
    stage_grade = ASC_GRADE_NAMES.get(int(digits))
    if not stage_grade:
        return None
    return stage_grade[0], stage_grade[1], match.group(2).strip()


def _parse_asctt_xml(
    content: bytes, shift_name: str
) -> list[tuple[str, list[str], list[tuple[int, dict[str, Any]]]]]:
    lowered = content[:16384].lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        fail("xml_unsafe_declaration", 415)
    declaration = re.search(br"<\?xml[^>]*encoding=[\"']([^\"']+)", content[:512], re.I)
    encoding = declaration.group(1).decode("ascii", "ignore").lower() if declaration else "utf-8"
    if encoding not in {"utf-8", "utf-8-sig", "windows-1256", "cp1256"}:
        fail("xml_encoding_not_supported", 415)
    try:
        root = ET.fromstring(content.decode(encoding))
    except (UnicodeDecodeError, ET.ParseError):
        fail("xml_structure_invalid", 415)
    if root.tag != "timetable" or not root.get("ascttversion"):
        fail("xml_not_asctt", 415)

    teachers = {item.get("id", ""): item for item in root.findall("./teachers/teacher")}
    subjects = {item.get("id", ""): item for item in root.findall("./subjects/subject")}
    classes = {item.get("id", ""): item for item in root.findall("./classes/class")}
    parsed_classes: dict[str, tuple[str, str, str]] = {}
    for class_id, item in classes.items():
        parsed = _asc_class_name(item.get("name", ""))
        if parsed:
            parsed_classes[class_id] = parsed

    teacher_subjects: dict[str, set[str]] = defaultdict(set)
    for lesson in root.findall("./lessons/lesson"):
        subject = subjects.get(lesson.get("subjectid", ""))
        if subject is None:
            continue
        for teacher_id in _asc_values(lesson.get("teacherid")):
            teacher_subjects[teacher_id].add(subject.get("name", "").strip())

    result: list[tuple[str, list[str], list[tuple[int, dict[str, Any]]]]] = []
    teacher_headers = ["كود المعلم", "اسم المعلم", "التخصص"]
    teacher_rows = [
        (
            index,
            {
                "كود المعلم": f"ASC-T-{teacher_id.lstrip('*')}",
                "اسم المعلم": item.get("name", "").strip(),
                "التخصص": "، ".join(sorted(teacher_subjects.get(teacher_id, set()))),
            },
        )
        for index, (teacher_id, item) in enumerate(teachers.items(), 1)
        if teacher_id and item.get("name", "").strip()
    ]
    result.append(("المعلمون من Timetables", teacher_headers, teacher_rows))

    subject_headers = ["رمز المادة", "اسم المادة"]
    subject_rows = [
        (
            index,
            {
                "رمز المادة": f"ASC-S-{subject_id.lstrip('*')}",
                "اسم المادة": item.get("name", "").strip(),
            },
        )
        for index, (subject_id, item) in enumerate(subjects.items(), 1)
        if subject_id and item.get("name", "").strip()
    ]
    result.append(("المواد من Timetables", subject_headers, subject_rows))

    structure_headers = ["رمز المرحلة", "المرحلة", "الصف", "الشعبة", "السعة"]
    structure_rows = [
        (
            index,
            {
                "رمز المرحلة": "ASC-PRIMARY" if stage == "المرحلة الابتدائية" else "ASC-INTERMEDIATE" if stage == "المرحلة المتوسطة" else "ASC-SECONDARY",
                "المرحلة": stage,
                "الصف": grade,
                "الشعبة": section,
                "السعة": "30",
            },
        )
        for index, (stage, grade, section) in enumerate(parsed_classes.values(), 1)
    ]
    result.append(("الصفوف والفصول من Timetables", structure_headers, structure_rows))

    offering_headers = ["المرحلة", "الصف", "الشعبة", "الفترة"]
    offering_rows = [
        (
            index,
            {"المرحلة": stage, "الصف": grade, "الشعبة": section, "الفترة": shift_name},
        )
        for index, (stage, grade, section) in enumerate(parsed_classes.values(), 1)
    ]
    result.append(("الفصول الدراسية من Timetables", offering_headers, offering_rows))

    assignment_headers = [
        "مفتاح المجموعة", "كود المعلم", "المادة", "المرحلة", "الصف", "الشعبة", "عدد الحصص"
    ]
    assignment_rows: list[tuple[int, dict[str, Any]]] = []
    for lesson_index, lesson in enumerate(root.findall("./lessons/lesson"), 1):
        lesson_id = lesson.get("id", str(lesson_index))
        subject = subjects.get(lesson.get("subjectid", ""))
        teacher_codes = [
            f"ASC-T-{teacher_id.lstrip('*')}"
            for teacher_id in _asc_values(lesson.get("teacherid"))
            if teacher_id in teachers
        ]
        try:
            weekly = int(float(lesson.get("periodsperweek", "0")))
        except ValueError:
            weekly = 0
        if subject is None or not teacher_codes or weekly < 1:
            continue
        class_ids = [item for item in _asc_values(lesson.get("classids")) if item in parsed_classes]
        for class_id in class_ids:
            stage, grade, section = parsed_classes[class_id]
            assignment_rows.append(
                (
                    len(assignment_rows) + 1,
                    {
                        "مفتاح المجموعة": f"ASC-L-{lesson_id.lstrip('*')}",
                        "كود المعلم": "|".join(teacher_codes),
                        "المادة": subject.get("name", "").strip(),
                        "المرحلة": stage,
                        "الصف": grade,
                        "الشعبة": section,
                        "عدد الحصص": str(weekly),
                    },
                )
            )
    result.append(("الإسنادات من Timetables", assignment_headers, assignment_rows))
    return result


def _unique_name_map(items: list[Any], attribute: str) -> dict[str, Any]:
    grouped: dict[str, list[Any]] = defaultdict(list)
    for item in items:
        grouped[normalize_text(getattr(item, attribute))].append(item)
    return {name: matches[0] for name, matches in grouped.items() if len(matches) == 1}


def _reconcile_asctt_existing(
    db: Session,
    tenant_id: uuid.UUID,
    school_id: uuid.UUID,
    sheets: list[tuple[str, list[str], list[tuple[int, dict[str, Any]]]]],
) -> None:
    teachers = list(
        db.scalars(
            select(Teacher)
            .join(TeacherSchoolMembership, TeacherSchoolMembership.teacher_id == Teacher.id)
            .where(
                Teacher.tenant_id == tenant_id,
                Teacher.is_active.is_(True),
                TeacherSchoolMembership.tenant_id == tenant_id,
                TeacherSchoolMembership.school_id == school_id,
                TeacherSchoolMembership.is_active.is_(True),
            )
        )
    )
    subjects = list(
        db.scalars(
            select(Subject).where(
                Subject.tenant_id == tenant_id,
                Subject.school_id == school_id,
                Subject.is_active.is_(True),
            )
        )
    )
    teachers_by_name = _unique_name_map(teachers, "name_ar")
    subjects_by_name = _unique_name_map(subjects, "name_ar")
    teacher_codes: dict[str, str] = {}
    for sheet_name, _, rows in sheets:
        if ASC_SHEET_ENTITIES.get(sheet_name) == "teachers":
            for _, values in rows:
                imported_code = str(values.get("كود المعلم", ""))
                existing = teachers_by_name.get(normalize_text(values.get("اسم المعلم", "")))
                if existing:
                    values["كود المعلم"] = existing.canonical_code
                    teacher_codes[imported_code] = existing.canonical_code
        elif ASC_SHEET_ENTITIES.get(sheet_name) == "subjects":
            for _, values in rows:
                existing = subjects_by_name.get(normalize_text(values.get("اسم المادة", "")))
                if existing:
                    values["رمز المادة"] = existing.code
    for sheet_name, _, rows in sheets:
        if ASC_SHEET_ENTITIES.get(sheet_name) != "assignments":
            continue
        for _, values in rows:
            values["كود المعلم"] = "|".join(
                teacher_codes.get(code, code)
                for code in str(values.get("كود المعلم", "")).split("|")
                if code
            )


def upload_job(
    db: Session,
    tenant_id: uuid.UUID,
    school_id: uuid.UUID,
    term_id: uuid.UUID | None,
    filename: str | None,
    content_type: str | None,
    content: bytes,
) -> ImportJob:
    school(db, tenant_id, school_id)
    if term_id and not db.scalar(
        select(Term.id)
        .join(AcademicYear, AcademicYear.id == Term.academic_year_id)
        .where(
            Term.id == term_id,
            Term.tenant_id == tenant_id,
            AcademicYear.tenant_id == tenant_id,
            AcademicYear.school_id == school_id,
        )
    ):
        fail("term_not_in_school")
    if len(content) > settings.import_max_file_bytes:
        fail("file_size_limit", 413)
    safe_name = _safe_filename(filename)
    extension = PurePath(safe_name).suffix.lower()
    if extension not in {".csv", ".xlsx", ".xml"}:
        fail("unsupported_file_type", 415)
    if extension == ".csv":
        sheets = _parse_csv(content)
    elif extension == ".xlsx":
        sheets = _parse_xlsx(content)
    else:
        shift = db.scalar(
            select(SchoolShift)
            .where(
                SchoolShift.tenant_id == tenant_id,
                SchoolShift.school_id == school_id,
                SchoolShift.is_active.is_(True),
            )
            .order_by(SchoolShift.order)
        )
        if not shift:
            fail("school_shift_required")
        sheets = _parse_asctt_xml(content, shift.name_ar)
        _reconcile_asctt_existing(db, tenant_id, school_id, sheets)
    total_rows = sum(len(rows) for _, _, rows in sheets)
    if total_rows > settings.import_max_rows:
        fail("row_limit_exceeded", 413)
    digest = hashlib.sha256(content).hexdigest()
    duplicate = bool(
        db.scalar(
            select(ImportJob.id).where(
                ImportJob.tenant_id == tenant_id,
                ImportJob.school_id == school_id,
                ImportJob.term_id == term_id,
                ImportJob.file_sha256 == digest,
                ImportJob.status == "committed",
            )
        )
    )
    detection = []
    job = ImportJob(
        tenant_id=tenant_id,
        school_id=school_id,
        term_id=term_id,
        source_filename=safe_name,
        content_type=(content_type or "application/octet-stream")[:120],
        file_size=len(content),
        file_sha256=digest,
        status="uploaded",
        detected_sheets=[],
        mapping={},
        validation_summary={},
        result_summary={},
        duplicate_file_warning=duplicate,
    )
    db.add(job)
    db.flush()
    for sheet_name, headers, rows in sheets:
        mapping = _suggest_mapping(headers)
        if extension == ".xml":
            entity = ASC_SHEET_ENTITIES[sheet_name]
            confidence = 1.0
        else:
            entity, confidence = _detect_entity(mapping)
        detection.append(
            {"name": sheet_name, "headers": headers, "entity_type": entity, "confidence": confidence, "suggested_mapping": mapping, "row_count": len(rows)}
        )
        for row_number, values in rows:
            db.add(
                ImportRow(
                    tenant_id=tenant_id,
                    import_job_id=job.id,
                    sheet_name=sheet_name,
                    source_row_number=row_number,
                    entity_type=entity,
                    source_values=values,
                    normalized_values={},
                    proposed_action="warning",
                    diagnostics=[],
                    before_values={},
                    after_values={},
                    excluded=False,
                )
            )
    job.detected_sheets = detection
    db.commit()
    return job


def _job(db: Session, tenant_id: uuid.UUID, school_id: uuid.UUID, job_id: uuid.UUID) -> ImportJob:
    job = db.scalar(
        select(ImportJob).where(
            ImportJob.id == job_id,
            ImportJob.tenant_id == tenant_id,
            ImportJob.school_id == school_id,
        )
    )
    if job is None:
        fail("import_job_not_found", 404)
    return job


def job_rows(db: Session, job: ImportJob) -> list[ImportRow]:
    return list(
        db.scalars(
            select(ImportRow)
            .where(ImportRow.import_job_id == job.id, ImportRow.tenant_id == job.tenant_id)
            .order_by(ImportRow.sheet_name, ImportRow.source_row_number)
        )
    )


def save_mapping(
    db: Session,
    tenant_id: uuid.UUID,
    school_id: uuid.UUID,
    job_id: uuid.UUID,
    mapping: dict[str, Any],
    allow_updates: bool,
) -> ImportJob:
    job = _job(db, tenant_id, school_id, job_id)
    if job.status == "committed":
        fail("import_already_committed", 409)
    known_sheets = {item["name"] for item in job.detected_sheets}
    if set(mapping) - known_sheets:
        fail("unknown_import_sheet")
    for row in job_rows(db, job):
        config = mapping.get(row.sheet_name)
        if not config or config["entity_type"] == "skip":
            row.excluded = True
            row.entity_type = "skip"
            row.normalized_values = {}
            continue
        row.excluded = False
        row.entity_type = config["entity_type"]
        row.normalized_values = {
            target: row.source_values.get(source, "")
            for source, target in config["columns"].items()
            if target
        }
        row.group_key = _plain(row.normalized_values.get("group_key")) or None
    job.mapping = {"sheets": mapping, "allow_updates": allow_updates}
    job.status = "mapped"
    db.commit()
    return job


def _plain(value: object) -> str:
    return "" if isinstance(value, dict) else str(value or "").strip()[:1000]


def _diag(row: ImportRow, severity: str, code: str, field: str | None = None) -> dict[str, Any]:
    messages = {
        "missing_required_field": "حقل مطلوب مفقود.",
        "formula_not_allowed": "لا تُقبل الصيغ في الحقول المستوردة.",
        "duplicate_row": "هذا الصف مكرر داخل الملف.",
        "teacher_ambiguous": "اسم المعلم يطابق أكثر من هوية؛ يلزم تحديد الكود.",
        "teacher_archived": "هوية المعلم مؤرشفة ولا تُفعّل بصمت.",
        "membership_inactive": "ارتباط المعلم بهذه المدرسة غير نشط.",
        "reference_not_found": "تعذر العثور على مرجع مطابق داخل المدرسة.",
        "existing_data_conflict": "توجد بيانات حالية ولن تُستبدل في الوضع الآمن.",
        "existing_data_difference": "تختلف القيم الواردة عن البيانات الحالية ولن تُستبدل في الوضع الآمن.",
        "invalid_weekly_count": "عدد الحصص الأسبوعية غير صالح.",
        "staged_reference_ambiguous": "يوجد أكثر من مرجع مرحلي مطابق داخل الملف.",
        "group_scalar_mismatch": "صفوف مجموعة التدريس لا تتفق على القيم الأساسية.",
    }
    return {"sheet": row.sheet_name, "row": row.source_row_number, "field": field, "severity": severity, "code": code, "message_ar": messages.get(code, code), "resolution_ar": "راجع المطابقة أو قيمة الخلية."}


def _resolve_structure(db: Session, job: ImportJob, values: dict[str, Any]) -> tuple[Stage | None, Grade | None, Section | None]:
    stage = db.scalar(select(Stage).where(Stage.tenant_id == job.tenant_id, Stage.school_id == job.school_id, func.lower(Stage.name_ar) == _plain(values.get("stage_name")).lower()))
    grade = db.scalar(select(Grade).where(Grade.tenant_id == job.tenant_id, Grade.stage_id == stage.id, func.lower(Grade.name_ar) == _plain(values.get("grade_name")).lower())) if stage else None
    section = db.scalar(select(Section).where(Section.tenant_id == job.tenant_id, Section.grade_id == grade.id, func.lower(Section.name_ar) == _plain(values.get("section_name")).lower())) if grade else None
    return stage, grade, section


def _key(*values: object) -> tuple[str, ...]:
    return tuple(normalize_text(_plain(value)) for value in values)


def _subject_by_name(
    db: Session, tenant_id: uuid.UUID, school_id: uuid.UUID, name: object
) -> Subject | None:
    normalized_name = normalize_text(name)
    return next(
        (
            subject
            for subject in db.scalars(
                select(Subject).where(
                    Subject.tenant_id == tenant_id,
                    Subject.school_id == school_id,
                )
            )
            if normalize_text(subject.name_ar) == normalized_name
        ),
        None,
    )


def _staged_indexes(rows: list[ImportRow]) -> dict[str, dict[tuple[str, ...], list[ImportRow]]]:
    indexes: dict[str, dict[tuple[str, ...], list[ImportRow]]] = {
        name: defaultdict(list) for name in ("stage", "grade", "section", "teacher", "subject_code", "subject_name", "resource_code", "resource_name", "offering", "curriculum")
    }
    for row in rows:
        if row.excluded:
            continue
        values = row.normalized_values
        if row.entity_type == "structure":
            indexes["stage"][_key(values.get("stage_name"))].append(row)
            indexes["grade"][_key(values.get("stage_name"), values.get("grade_name"))].append(row)
            indexes["section"][_key(values.get("stage_name"), values.get("grade_name"), values.get("section_name"))].append(row)
        elif row.entity_type == "teachers":
            indexes["teacher"][_key(values.get("teacher_code"))].append(row)
        elif row.entity_type == "subjects":
            indexes["subject_code"][_key(values.get("subject_code"))].append(row)
            indexes["subject_name"][_key(values.get("subject_name"))].append(row)
        elif row.entity_type == "resources":
            indexes["resource_code"][_key(values.get("resource_code"))].append(row)
            indexes["resource_name"][_key(values.get("resource_name"))].append(row)
        elif row.entity_type == "offerings":
            indexes["offering"][_key(values.get("stage_name"), values.get("grade_name"), values.get("section_name"))].append(row)
        elif row.entity_type == "curriculum":
            indexes["curriculum"][_key(values.get("stage_name"), values.get("grade_name"), values.get("subject_name"))].append(row)
    return indexes


def _one_staged(index: dict[tuple[str, ...], list[ImportRow]], key: tuple[str, ...]) -> ImportRow | None:
    matches = index.get(key, [])
    return matches[0] if len(matches) == 1 else None


def _add_difference(row: ImportRow, before: dict[str, Any], after: dict[str, Any], allow_updates: bool) -> str:
    differences = {field: value for field, value in after.items() if before.get(field) != value}
    row.before_values = before
    row.after_values = {**before, **differences}
    if not differences:
        return "skip_unchanged"
    if allow_updates:
        return "update"
    row.diagnostics.append(_diag(row, "warning", "existing_data_difference"))
    return "warning"


def _staged_assignment_projection(
    overlay: dict[str, dict[tuple[str, ...], list[ImportRow]]],
    section_keys: set[tuple[str, ...]],
    teacher_codes: set[str],
    subject_name: object,
    count: int,
) -> dict[str, Any]:
    coverage: list[dict[str, Any]] = []
    for section_key in sorted(section_keys):
        curriculum_key = (section_key[0], section_key[1], normalize_text(subject_name))
        required = next(
            (
                int(_plain(item.normalized_values.get("weekly_occurrences")))
                for item in overlay["curriculum"].get(curriculum_key, [])
                if _plain(item.normalized_values.get("weekly_occurrences"))
            ),
            0,
        )
        coverage.append(
            {
                "section": list(section_key),
                "required": required,
                "current_assigned": 0,
                "delta": count,
                "projected_assigned": count,
                "projected_status": "over" if required and count > required else "complete" if required == count else "partial",
            }
        )
    return {
        "coverage": coverage,
        "teacher_workloads": [
            {"teacher_code": code, "current_workload": 0, "delta": count, "projected_workload": count, "limit": 24, "exceeds_limit": count > 24}
            for code in sorted(teacher_codes)
        ],
        "warnings": [],
    }


def validate_job(db: Session, tenant_id: uuid.UUID, school_id: uuid.UUID, job_id: uuid.UUID) -> ImportJob:
    job = _job(db, tenant_id, school_id, job_id)
    if job.status == "committed":
        fail("import_already_committed", 409)
    if not job.mapping:
        fail("mapping_required")
    rows = job_rows(db, job)
    overlay = _staged_indexes(rows)
    allow_updates = bool(job.mapping.get("allow_updates"))
    seen: set[tuple[str, str]] = set()
    counts: Counter[str] = Counter()
    assignment_groups: dict[str, list[ImportRow]] = defaultdict(list)
    for row in rows:
        row.diagnostics = []
        row.before_values = {}
        row.after_values = dict(row.normalized_values)
        if row.excluded or row.entity_type == "skip":
            row.proposed_action = "skip_unchanged"
            counts["skipped"] += 1
            continue
        values = row.normalized_values
        for field in REQUIRED_FIELDS[row.entity_type]:
            if not _plain(values.get(field)):
                row.diagnostics.append(_diag(row, "error", "missing_required_field", field))
        for field, value in values.items():
            if isinstance(value, dict) and "formula" in value:
                row.diagnostics.append(_diag(row, "error", "formula_not_allowed", field))
        fingerprint = (row.entity_type, repr(sorted((key, _plain(value)) for key, value in values.items())))
        if fingerprint in seen:
            row.diagnostics.append(_diag(row, "error", "duplicate_row"))
        seen.add(fingerprint)
        action = "create"
        if row.entity_type == "teachers":
            code = _plain(values.get("teacher_code"))
            staged_matches = overlay["teacher"].get(_key(code), [])
            if len(staged_matches) > 1:
                row.diagnostics.append(_diag(row, "error", "staged_reference_ambiguous", "teacher_code"))
            teacher = db.scalar(select(Teacher).where(Teacher.tenant_id == tenant_id, Teacher.canonical_code == code))
            if teacher:
                membership = db.scalar(select(TeacherSchoolMembership).where(TeacherSchoolMembership.tenant_id == tenant_id, TeacherSchoolMembership.school_id == school_id, TeacherSchoolMembership.teacher_id == teacher.id))
                if not teacher.is_active:
                    row.diagnostics.append(_diag(row, "error", "teacher_archived", "teacher_code"))
                elif membership and not membership.is_active:
                    row.diagnostics.append(_diag(row, "error", "membership_inactive", "teacher_code"))
                if membership:
                    action = _add_difference(row, {"name_ar": teacher.name_ar, "specialty": teacher.specialty_reference or ""}, {"name_ar": _plain(values.get("teacher_name")), "specialty": _plain(values.get("specialty"))}, allow_updates)
                    row.before_values["teacher_id"] = str(teacher.id)
                else:
                    action = "link_existing"
                    row.before_values = {"teacher_id": str(teacher.id), "name_ar": teacher.name_ar}
            elif not code:
                matches = list(db.scalars(select(Teacher).where(Teacher.tenant_id == tenant_id, func.lower(Teacher.name_ar) == _plain(values.get("teacher_name")).lower())))
                if len(matches) != 1:
                    row.diagnostics.append(_diag(row, "error", "teacher_ambiguous", "teacher_name"))
        elif row.entity_type == "subjects":
            subject_existing = db.scalar(select(Subject).where(Subject.tenant_id == tenant_id, Subject.school_id == school_id, Subject.code == _plain(values.get("subject_code"))))
            if len(overlay["subject_code"].get(_key(values.get("subject_code")), [])) > 1:
                row.diagnostics.append(_diag(row, "error", "staged_reference_ambiguous", "subject_code"))
            if subject_existing:
                action = _add_difference(row, {"name_ar": subject_existing.name_ar}, {"name_ar": _plain(values.get("subject_name"))}, allow_updates)
        elif row.entity_type == "resources":
            resource_existing = db.scalar(select(Resource).where(Resource.tenant_id == tenant_id, Resource.school_id == school_id, Resource.code == _plain(values.get("resource_code"))))
            if len(overlay["resource_code"].get(_key(values.get("resource_code")), [])) > 1:
                row.diagnostics.append(_diag(row, "error", "staged_reference_ambiguous", "resource_code"))
            if resource_existing:
                action = _add_difference(row, {"name_ar": resource_existing.name_ar, "resource_type": resource_existing.resource_type, "capacity": resource_existing.capacity}, {"name_ar": _plain(values.get("resource_name")), "resource_type": _plain(values.get("resource_type")) or "room", "capacity": int(_plain(values.get("capacity")) or 30)}, allow_updates)
        elif row.entity_type == "structure":
            _, _, section = _resolve_structure(db, job, values)
            action = "skip_unchanged" if section else "create"
        elif row.entity_type in {"curriculum", "offerings", "assignments"}:
            _, grade, section = _resolve_structure(db, job, values)
            subject = (
                _subject_by_name(db, tenant_id, school_id, values.get("subject_name"))
                if row.entity_type != "offerings"
                else None
            )
            grade_matches = overlay["grade"].get(_key(values.get("stage_name"), values.get("grade_name")), [])
            grade_staged = grade_matches[0] if grade_matches else None
            section_staged = _one_staged(overlay["section"], _key(values.get("stage_name"), values.get("grade_name"), values.get("section_name")))
            subject_staged = _one_staged(overlay["subject_name"], _key(values.get("subject_name")))
            ambiguous = ((row.entity_type != "curriculum" and not section and len(overlay["section"].get(_key(values.get("stage_name"), values.get("grade_name"), values.get("section_name")), [])) > 1) or (row.entity_type != "offerings" and not subject and len(overlay["subject_name"].get(_key(values.get("subject_name")), [])) > 1))
            if ambiguous:
                row.diagnostics.append(_diag(row, "error", "staged_reference_ambiguous"))
            elif (not grade and not grade_staged) or (row.entity_type != "curriculum" and not section and not section_staged) or (row.entity_type != "offerings" and not subject and not subject_staged):
                row.diagnostics.append(_diag(row, "error", "reference_not_found"))
            try:
                count = int(_plain(values.get("weekly_occurrences"))) if row.entity_type != "offerings" else 1
                if count < 1 or count > 60:
                    raise ValueError
            except ValueError:
                row.diagnostics.append(_diag(row, "error", "invalid_weekly_count", "weekly_occurrences"))
            if row.entity_type == "curriculum" and grade and subject:
                requirement_existing = db.scalar(select(CurriculumRequirement).where(CurriculumRequirement.tenant_id == tenant_id, CurriculumRequirement.school_id == school_id, CurriculumRequirement.grade_id == grade.id, CurriculumRequirement.subject_id == subject.id))
                if requirement_existing:
                    row.before_values = {"weekly_occurrences": requirement_existing.weekly_occurrences}
                    row.after_values = {"weekly_occurrences": count}
                    if allow_updates:
                        action = "update"
                    else:
                        action = "conflict"
                        row.diagnostics.append(_diag(row, "error", "existing_data_conflict"))
            if row.entity_type == "offerings" and not job.term_id:
                row.diagnostics.append(_diag(row, "error", "reference_not_found", "term"))
            if row.entity_type == "assignments":
                assignment_groups[row.group_key or str(row.id)].append(row)
        if any(item["severity"] == "error" for item in row.diagnostics):
            action = "conflict"
            counts["errors"] += 1
        counts["warnings"] += sum(item["severity"] == "warning" for item in row.diagnostics)
        counts[action] += 1
        row.proposed_action = action
        # JSON mutations must be reassigned so SQLAlchemy persists row diagnostics.
        row.diagnostics = [dict(item) for item in row.diagnostics]
        flag_modified(row, "diagnostics")
    for group_key, grouped in assignment_groups.items():
        scalars = {(_key(row.normalized_values.get("subject_name")), _plain(row.normalized_values.get("weekly_occurrences")), str(job.term_id)) for row in grouped}
        if len(scalars) != 1:
            for row in grouped:
                row.diagnostics.append(_diag(row, "error", "group_scalar_mismatch", "group_key"))
                row.proposed_action = "conflict"
            continue
        teacher_ids: set[uuid.UUID] = set()
        offering_ids: set[uuid.UUID] = set()
        resource_ids: set[uuid.UUID] = set()
        staged_teachers: set[str] = set()
        staged_sections: set[tuple[str, ...]] = set()
        staged_resources: set[str] = set()
        subject_obj: Subject | None = None
        failed_group = False
        for row in grouped:
            values = row.normalized_values
            subject_obj = subject_obj or _subject_by_name(
                db, tenant_id, school_id, values.get("subject_name")
            )
            for code in {item.strip() for item in _plain(values.get("teacher_code")).split("|") if item.strip()}:
                teacher = db.scalar(select(Teacher).join(TeacherSchoolMembership, TeacherSchoolMembership.teacher_id == Teacher.id).where(Teacher.tenant_id == tenant_id, Teacher.canonical_code == code, Teacher.is_active.is_(True), TeacherSchoolMembership.school_id == school_id, TeacherSchoolMembership.is_active.is_(True)))
                staged = _one_staged(overlay["teacher"], _key(code))
                if teacher:
                    teacher_ids.add(teacher.id)
                elif staged:
                    staged_teachers.add(code)
                else:
                    row.diagnostics.append(_diag(row, "error", "reference_not_found", "teacher_code"))
                    failed_group = True
            _, _, section = _resolve_structure(db, job, values)
            section_key = _key(values.get("stage_name"), values.get("grade_name"), values.get("section_name"))
            offering = db.scalar(select(SectionOffering).where(SectionOffering.tenant_id == tenant_id, SectionOffering.school_id == school_id, SectionOffering.term_id == job.term_id, SectionOffering.section_id == section.id, SectionOffering.is_active.is_(True))) if section and job.term_id else None
            if offering:
                offering_ids.add(offering.id)
            elif _one_staged(overlay["offering"], section_key):
                staged_sections.add(section_key)
            else:
                row.diagnostics.append(_diag(row, "error", "reference_not_found", "section_name"))
                failed_group = True
            resource_code = _plain(values.get("resource_code"))
            if resource_code:
                resource = db.scalar(select(Resource).where(Resource.tenant_id == tenant_id, Resource.school_id == school_id, Resource.code == resource_code, Resource.is_active.is_(True)))
                if resource:
                    resource_ids.add(resource.id)
                elif _one_staged(overlay["resource_code"], _key(resource_code)):
                    staged_resources.add(resource_code)
                else:
                    row.diagnostics.append(_diag(row, "error", "reference_not_found", "resource_code"))
                    failed_group = True
        if not subject_obj and not _one_staged(overlay["subject_name"], next(iter(scalars))[0]):
            failed_group = True
        if failed_group:
            for row in grouped:
                row.proposed_action = "conflict"
            continue
        count = int(next(iter(scalars))[1])
        if subject_obj and not staged_teachers and not staged_sections and not staged_resources:
            preview = preview_assignment(db, tenant_id, school_id, {"term_id": job.term_id, "subject_id": subject_obj.id, "weekly_occurrences": count, "teacher_ids": list(teacher_ids), "section_offering_ids": list(offering_ids), "resource_ids": list(resource_ids)})
            preview_data = preview.model_dump(mode="json")
        else:
            preview_data = _staged_assignment_projection(overlay, staged_sections, staged_teachers, grouped[0].normalized_values.get("subject_name"), count)
        for row in grouped:
            row.after_values = {**row.after_values, "assignment_preview": preview_data, "aggregate_group_key": group_key}
            flag_modified(row, "after_values")
            flag_modified(row, "diagnostics")
    counts = Counter()
    for row in rows:
        if row.excluded or row.entity_type == "skip":
            counts["skipped"] += 1
        else:
            if any(item["severity"] == "error" for item in row.diagnostics):
                row.proposed_action = "conflict"
                counts["errors"] += 1
            counts["warnings"] += sum(item["severity"] == "warning" for item in row.diagnostics)
            counts[row.proposed_action] += 1
        row.diagnostics = [dict(item) for item in row.diagnostics]
        flag_modified(row, "diagnostics")
    total = len(job_rows(db, job))
    job.validation_summary = {"total": total, "valid": total - counts["errors"] - counts["skipped"], "errors": counts["errors"], "warnings": counts["warnings"], "actions": dict(counts)}
    job.validated_at = datetime.now(timezone.utc)
    job.status = "ready" if counts["errors"] == 0 else "validated"
    db.commit()
    return job


def exclude_rows(db: Session, tenant_id: uuid.UUID, school_id: uuid.UUID, job_id: uuid.UUID, row_ids: list[uuid.UUID]) -> ImportJob:
    job = _job(db, tenant_id, school_id, job_id)
    rows = list(db.scalars(select(ImportRow).where(ImportRow.import_job_id == job.id, ImportRow.tenant_id == tenant_id, ImportRow.id.in_(row_ids))))
    if len(rows) != len(row_ids):
        fail("import_row_not_in_job")
    for row in rows:
        row.excluded = True
    job.status = "mapped"
    db.commit()
    return job


def _find_structure(db: Session, job: ImportJob, values: dict[str, Any]) -> tuple[Stage, Grade, Section]:
    stage, grade, section = _resolve_structure(db, job, values)
    if not stage or not grade or not section:
        fail("reference_not_found")
    return stage, grade, section


def _matching_current_assignments(
    db: Session,
    tenant_id: uuid.UUID,
    school_id: uuid.UUID,
    term_id: uuid.UUID,
    subject_id: uuid.UUID,
    offering_ids: set[uuid.UUID],
) -> list[TeachingAssignment]:
    target_subject_name = db.scalar(
        select(Subject.name_ar).where(
            Subject.id == subject_id,
            Subject.tenant_id == tenant_id,
            Subject.school_id == school_id,
        )
    )
    target_sections = {
        _key(stage_name, grade_name, section_name)
        for stage_name, grade_name, section_name in db.execute(
            select(Stage.name_ar, Grade.name_ar, Section.name_ar)
            .join(Grade, Grade.stage_id == Stage.id)
            .join(Section, Section.grade_id == Grade.id)
            .join(SectionOffering, SectionOffering.section_id == Section.id)
            .where(
                SectionOffering.tenant_id == tenant_id,
                SectionOffering.id.in_(offering_ids),
            )
        )
    }
    candidates = list(
        db.scalars(
            select(TeachingAssignment)
            .where(
                TeachingAssignment.tenant_id == tenant_id,
                TeachingAssignment.school_id == school_id,
                TeachingAssignment.term_id == term_id,
            )
            .order_by(TeachingAssignment.created_at, TeachingAssignment.id)
        )
    )
    matches: list[TeachingAssignment] = []
    for assignment in candidates:
        has_teacher = db.scalar(
            select(TeachingAssignmentTeacher.id).where(
                TeachingAssignmentTeacher.tenant_id == tenant_id,
                TeachingAssignmentTeacher.teaching_assignment_id == assignment.id,
            )
        )
        candidate_subject_name = db.scalar(
            select(Subject.name_ar).where(
                Subject.id == assignment.subject_id,
                Subject.tenant_id == tenant_id,
            )
        )
        current_sections = {
            _key(stage_name, grade_name, section_name)
            for stage_name, grade_name, section_name in db.execute(
                select(Stage.name_ar, Grade.name_ar, Section.name_ar)
                .join(Grade, Grade.stage_id == Stage.id)
                .join(Section, Section.grade_id == Grade.id)
                .join(SectionOffering, SectionOffering.section_id == Section.id)
                .join(
                    TeachingAssignmentSection,
                    TeachingAssignmentSection.section_offering_id == SectionOffering.id,
                )
                .where(
                    TeachingAssignmentSection.tenant_id == tenant_id,
                    TeachingAssignmentSection.teaching_assignment_id == assignment.id,
                )
            )
        }
        if (
            has_teacher
            and normalize_text(candidate_subject_name) == normalize_text(target_subject_name)
            and current_sections == target_sections
        ):
            matches.append(assignment)
    return matches


def _retire_duplicate_assignment(
    db: Session, tenant_id: uuid.UUID, assignment: TeachingAssignment
) -> None:
    has_history = bool(
        db.scalar(
            select(TimetableEntry.id).where(
                TimetableEntry.tenant_id == tenant_id,
                TimetableEntry.assignment_id == assignment.id,
            )
        )
        or db.scalar(
            select(WorkingTimetableEntry.id).where(
                WorkingTimetableEntry.tenant_id == tenant_id,
                WorkingTimetableEntry.assignment_id == assignment.id,
            )
        )
    )
    if has_history:
        db.execute(
            delete(TeachingAssignmentTeacher).where(
                TeachingAssignmentTeacher.tenant_id == tenant_id,
                TeachingAssignmentTeacher.teaching_assignment_id == assignment.id,
            )
        )
    else:
        db.delete(assignment)


def commit_job(db: Session, tenant_id: uuid.UUID, school_id: uuid.UUID, job_id: uuid.UUID, acknowledge_warnings: bool) -> ImportJob:
    job = _job(db, tenant_id, school_id, job_id)
    if job.status == "committed" or job.committed_at:
        fail("import_already_committed", 409)
    if job.status != "ready":
        fail("import_not_ready", 409)
    if job.validation_summary.get("warnings") and not acknowledge_warnings:
        fail("warnings_acknowledgement_required", 409)
    rows = [row for row in job_rows(db, job) if not row.excluded and row.proposed_action in {"create", "link_existing", "update"}]
    counts: Counter[str] = Counter()
    try:
        for row in [item for item in rows if item.entity_type == "structure"]:
            values = row.normalized_values
            stage, grade, section = _resolve_structure(db, job, values)
            if not stage:
                stage = save_resource(db, tenant_id, school_id, "stages", {"code": _plain(values.get("stage_code")) or f"ST-{row.source_row_number}", "name_ar": _plain(values.get("stage_name")), "order": row.source_row_number}, commit_changes=False)
            if not grade:
                grade = save_resource(db, tenant_id, school_id, "grades", {"stage_id": stage.id, "name_ar": _plain(values.get("grade_name")), "order": row.source_row_number}, commit_changes=False)
            if not section:
                save_resource(db, tenant_id, school_id, "sections", {"grade_id": grade.id, "name_ar": _plain(values.get("section_name")), "capacity": int(_plain(values.get("capacity")) or 30)}, commit_changes=False)
            counts["created"] += 1
        for row in [item for item in rows if item.entity_type == "teachers"]:
            values = row.normalized_values
            teacher = db.scalar(select(Teacher).where(Teacher.tenant_id == tenant_id, Teacher.canonical_code == _plain(values.get("teacher_code"))))
            if teacher:
                membership = db.scalar(select(TeacherSchoolMembership).where(TeacherSchoolMembership.tenant_id == tenant_id, TeacherSchoolMembership.school_id == school_id, TeacherSchoolMembership.teacher_id == teacher.id))
                if row.proposed_action == "update" and membership:
                    update_teacher(db, tenant_id, school_id, teacher.id, {"canonical_code": teacher.canonical_code, "name_ar": _plain(values.get("teacher_name")), "name_en": teacher.name_en, "specialty_reference": _plain(values.get("specialty")) or None, "base_workload": teacher.base_workload, "teaching_workload_limit": teacher.teaching_workload_limit, "is_active": teacher.is_active}, commit_changes=False)
                    counts["updated"] += 1
                elif not membership:
                    link_teacher(db, tenant_id, school_id, {"teacher_id": teacher.id, "local_employee_code": _plain(values.get("teacher_code")), "is_home_school": False, "is_active": True}, commit_changes=False)
                    counts["linked"] += 1
            else:
                create_teacher(db, tenant_id, school_id, {"canonical_code": _plain(values.get("teacher_code")), "name_ar": _plain(values.get("teacher_name")), "specialty_reference": _plain(values.get("specialty")) or None, "base_workload": 0, "teaching_workload_limit": 24, "is_active": True, "local_employee_code": _plain(values.get("teacher_code")), "is_home_school": False}, commit_changes=False)
                counts["created"] += 1
        for entity, kind in (("subjects", "subjects"), ("resources", "resources")):
            for row in [item for item in rows if item.entity_type == entity]:
                values = row.normalized_values
                payload: dict[str, Any] = ({"code": _plain(values.get("subject_code")), "name_ar": _plain(values.get("subject_name")), "name_en": None, "is_active": True} if entity == "subjects" else {"code": _plain(values.get("resource_code")), "name_ar": _plain(values.get("resource_name")), "resource_type": _plain(values.get("resource_type")) or "room", "capacity": int(_plain(values.get("capacity")) or 30), "exclusive": True, "is_active": True})
                existing_entity = db.scalar(select(Subject).where(Subject.tenant_id == tenant_id, Subject.school_id == school_id, Subject.code == _plain(values.get("subject_code")))) if entity == "subjects" else db.scalar(select(Resource).where(Resource.tenant_id == tenant_id, Resource.school_id == school_id, Resource.code == _plain(values.get("resource_code"))))
                if isinstance(existing_entity, Subject):
                    payload.update(name_en=existing_entity.name_en, is_active=existing_entity.is_active)
                elif isinstance(existing_entity, Resource):
                    payload.update(exclusive=existing_entity.exclusive, is_active=existing_entity.is_active)
                save_catalog(db, tenant_id, school_id, kind, payload, existing_entity.id if row.proposed_action == "update" and existing_entity else None, commit_changes=False)
                counts["updated" if existing_entity else "created"] += 1
        for row in [item for item in rows if item.entity_type == "curriculum"]:
            values = row.normalized_values
            _, grade, _ = _resolve_structure(db, job, values)
            subject = _subject_by_name(db, tenant_id, school_id, values.get("subject_name"))
            if not grade or not subject:
                fail("reference_not_found")
            existing = db.scalar(select(CurriculumRequirement).where(CurriculumRequirement.tenant_id == tenant_id, CurriculumRequirement.school_id == school_id, CurriculumRequirement.grade_id == grade.id, CurriculumRequirement.subject_id == subject.id))
            save_catalog(db, tenant_id, school_id, "requirements", {"grade_id": grade.id, "subject_id": subject.id, "weekly_occurrences": int(_plain(values.get("weekly_occurrences"))), "notes": None}, existing.id if existing else None, commit_changes=False)
            counts["updated" if existing else "created"] += 1
        for row in [item for item in rows if item.entity_type == "offerings"]:
            _, _, section = _find_structure(db, job, row.normalized_values)
            shift = db.scalar(select(SchoolShift).where(SchoolShift.tenant_id == tenant_id, SchoolShift.school_id == school_id, func.lower(SchoolShift.name_ar) == _plain(row.normalized_values.get("shift_name")).lower(), SchoolShift.is_active.is_(True)))
            if not shift or not job.term_id:
                fail("reference_not_found")
            save_offerings(db, tenant_id, school_id, {"offerings": [{"term_id": job.term_id, "section_id": section.id, "shift_id": shift.id, "is_active": True}]}, commit_changes=False)
            counts["created"] += 1
        assignment_rows = [item for item in rows if item.entity_type == "assignments"]
        groups: dict[str, list[ImportRow]] = defaultdict(list)
        for row in assignment_rows:
            groups[row.group_key or str(row.id)].append(row)
        for grouped in groups.values():
            first = grouped[0]
            values = first.normalized_values
            subject = _subject_by_name(db, tenant_id, school_id, values.get("subject_name"))
            teachers: set[uuid.UUID] = set()
            offerings: set[uuid.UUID] = set()
            resources: set[uuid.UUID] = set()
            for row in grouped:
                _, _, section = _find_structure(db, job, row.normalized_values)
                offering = db.scalar(select(SectionOffering).where(SectionOffering.tenant_id == tenant_id, SectionOffering.school_id == school_id, SectionOffering.term_id == job.term_id, SectionOffering.section_id == section.id, SectionOffering.is_active.is_(True)))
                if not offering:
                    fail("reference_not_found")
                offerings.add(offering.id)
                codes = [_plain(item) for item in _plain(row.normalized_values.get("teacher_code")).split("|")]
                teachers.update(db.scalars(select(Teacher.id).join(TeacherSchoolMembership, TeacherSchoolMembership.teacher_id == Teacher.id).where(Teacher.tenant_id == tenant_id, Teacher.canonical_code.in_(codes), Teacher.is_active.is_(True), TeacherSchoolMembership.school_id == school_id, TeacherSchoolMembership.is_active.is_(True))))
                resource_code = _plain(row.normalized_values.get("resource_code"))
                if resource_code:
                    resource = db.scalar(select(Resource).where(Resource.tenant_id == tenant_id, Resource.school_id == school_id, Resource.code == resource_code, Resource.is_active.is_(True)))
                    if not resource:
                        fail("reference_not_found")
                    resources.add(resource.id)
            if not subject or not job.term_id or not teachers:
                fail("reference_not_found")
            existing_assignments = _matching_current_assignments(
                db,
                tenant_id,
                school_id,
                job.term_id,
                subject.id,
                offerings,
            )
            canonical = existing_assignments[0] if existing_assignments else None
            save_assignment(
                db,
                tenant_id,
                school_id,
                {
                    "term_id": job.term_id,
                    "subject_id": subject.id,
                    "weekly_occurrences": int(_plain(values.get("weekly_occurrences"))),
                    "teacher_ids": list(teachers),
                    "section_offering_ids": list(offerings),
                    "resource_ids": list(resources),
                    "notes": f"Import {job.id}",
                },
                assignment_id=canonical.id if canonical else None,
                commit_changes=False,
            )
            for duplicate in existing_assignments[1:]:
                _retire_duplicate_assignment(db, tenant_id, duplicate)
            counts["updated" if canonical else "created"] += 1
        job.status = "committed"
        job.committed_at = datetime.now(timezone.utc)
        job.result_summary = dict(counts)
        db.commit()
    except Exception as exc:
        db.rollback()
        failed = _job(db, tenant_id, school_id, job_id)
        failed.status = "failed"
        failed.result_summary = {"error": type(exc).__name__}
        db.commit()
        raise HTTPException(status_code=409, detail={"code": "atomic_import_failed"}) from exc
    return job


def serialize_job(db: Session, job: ImportJob) -> dict[str, Any]:
    return {"id": job.id, "school_id": job.school_id, "term_id": job.term_id, "source_filename": job.source_filename, "file_size": job.file_size, "file_sha256": job.file_sha256, "status": job.status, "detected_sheets": job.detected_sheets, "mapping": job.mapping, "validation_summary": job.validation_summary, "result_summary": job.result_summary, "duplicate_file_warning": job.duplicate_file_warning, "rows": job_rows(db, job)}


TEMPLATES = {
    "teachers": ["كود المعلم", "اسم المعلم", "التخصص"],
    "subjects": ["رمز المادة", "اسم المادة"],
    "structure": ["رمز المرحلة", "المرحلة", "الصف", "الشعبة", "السعة"],
    "curriculum": ["المرحلة", "الصف", "المادة", "عدد الحصص"],
    "resources": ["رمز المورد", "المورد", "نوع المورد", "السعة"],
    "assignments": ["كود المعلم", "المادة", "المرحلة", "الصف", "الشعبة", "عدد الحصص", "رمز المورد", "مفتاح المجموعة"],
}


def template_csv(kind: str) -> bytes:
    if kind not in TEMPLATES:
        fail("unknown_import_template", 404)
    stream = io.StringIO()
    csv.writer(stream).writerow(TEMPLATES[kind])
    return ("\ufeff" + stream.getvalue()).encode("utf-8")
